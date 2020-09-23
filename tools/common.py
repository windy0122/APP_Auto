import random
import logging
import requests
import json
import openpyxl
import time
from tools.project_path import *
from tools.test_is_exists import TestIsExists


class StartBefore(object):
    shop_tk = None
    shop_id = None
    TestIsExists().test_is_exists()
    # shop_customer = None

    # 登录获取店铺tk和shop_id
    # 将token写入init表中
    def login(self):
        url = 'https://uatleague.round-table-union.com/api/rts/base/home/login/V141'

        header = {'content-type': 'application/json',
                  'version': '1.4.1',
                  'platform': 'iOS'
                  }

        payload = {
                "smsCodeToken": None,
                "standbyUniqueIdentification": "_Android10",
                "deviceUniqueIdentification": "",
                "password": "e10adc3949ba59abbe56e057f20f883e",
                "loginType": "1",
                "smsCode": None,
                "loginCode": "13411111111",
                "source": "",
                "applicationId": "AjwG1Pb2_9CZbEZ3QnWIc1f7HRbPBXo6-ismtWk5Y27H"
            }

        try:
            r = requests.post(url, headers=header, data=json.dumps(payload), verify=False)
            res = r.json()
            shop_tk = res['val']['tk']
            self.write_back_init(test_tmp_path, 'init', 6, shop_tk)
        except Exception as e:
            logging.info('登录失败')
            raise e
    #
    # # 查询店铺顾客手机号
    # # @staticmethod
    # def get_phone_num(self):
    #
    #     url = 'https://uatleague.round-table-union.com/api/rts/member/membership/searchMembership/V133?condition='
    #
    #     header = {'content-type': 'application/json',
    #               'version': '1.4.1',
    #               'platform': 'iOS',
    #               'tk': self.get_token()[0]
    #               }
    #
    #     r = requests.get(url, headers=header, verify=False)
    #     # print(r.json())
    #     member_res = r.json()['val']['membershipList']
    #     self.shop_customer = []
    #     if len(member_res) > 0:
    #         for i in range(len(member_res)):
    #             self.shop_customer.append(member_res[i]['membershipMobile'])
    #
    #     # 返回顾客号码列表
    #     return self.shop_customer

    # 随机生成顾客手机号
    @staticmethod
    def create_phone():
        prelist = ['13', '14', '15', '16', '17', '18', '19']
        new_phone_num = random.choice(prelist) + ''.join(random.choice('0123456789') for i in range(9))
        return new_phone_num

    # 验证顾客号码是否是店铺会员
    # @staticmethod
    # def new_customer_phone(self):
    #     self.get_phone_num()
    #     phone_num = self.create_phone()
    #     if phone_num in self.shop_customer:
    #         self.create_phone()
    #     else:
    #         return phone_num

    # 获取init表中数据(1、顾客id   2、储值卡id   3、计次卡id    4、年卡id    5、员工id    6、店铺tk)
    @staticmethod
    def get_data_init(num):
        wb = openpyxl.load_workbook(test_tmp_path)
        sheet = wb['init']
        init_data = sheet.cell(2, num).value
        if init_data is None:
            init_data = '1'
        return init_data

    # 写入result(返回数据)，写入TestResult(pass or failed)
    @staticmethod
    def write_back(file_name, sheet_name, case_id, i, result, TestResult, current_path):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 1).value = case_id
        sheet.cell(i, 2).value = result
        sheet.cell(i, 3).value = TestResult
        sheet.cell(i, 4).value = current_path
        wb.save(file_name)

    # 写入数据到init表中
    @staticmethod
    def write_back_init(file_name, sheet_name, i, result):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(2, i).value = result
        wb.save(file_name)

    # 将顾客卡id写入到init表中
    def write_customer_card_id(self, res):
        if isinstance(res['val'], list):
            for i in range(len(res['val'])):
                if res['val'][i]['cardType'] == '1':
                    self.write_back_init(test_tmp_path, 'init', 2, res['val'][i]['membershipCardId'])
                elif res['val'][i]['cardType'] == '2':
                    self.write_back_init(test_tmp_path, 'init', 3, res['val'][i]['membershipCardId'])
                elif res['val'][i]['cardType'] == '3':
                    self.write_back_init(test_tmp_path, 'init', 4, res['val'][i]['membershipCardId'])
                else:
                    logging.info('顾客没有卡')

    # 将新建的员工id写入到init表中
    def write_employee_id(self, res):
        if isinstance(res['val'], list) and len(res['val']) > 1:
            self.write_back_init(test_tmp_path, 'init', 5, res['val'][1]['employeeId'])

    # 获取当前时间时间戳
    @staticmethod
    def get_time_now():
        time_now = round(time.time()*1000)
        return str(time_now)


if __name__ == '__main__':
    start = StartBefore()
    print(start.create_phone())
    # print(start.test_is_exists())
    # print(start.new_customer_phone())
    # print(start.get_data_init(6))
    # print(start.get_time_now())







