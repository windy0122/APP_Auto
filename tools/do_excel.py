import openpyxl
from tools.read_config import ReadConfig
from tools.project_path import *
from tools.common import StartBefore


class DoExcel(object):
    @staticmethod
    def get_data(file_name, sheet_name):
        # user_tk = StartBefore().get_token()[0]
        customer_phone = StartBefore().create_phone()
        url_uat = ReadConfig().get_config(test_config_path, 'URL', 'url')
        wb = openpyxl.load_workbook(file_name)
        sheet = wb[sheet_name]
        # mode = eval(ReadConfig.get_config(test_config_path, 'MODE', 'mode'))
        test_data = []
        # for key in mode:
        #     sheet = wb[key]
        #     if mode[key] == 'all':
        for i in range(2, sheet.max_row + 1):
            row_data = dict()
            row_data['case_id'] = sheet.cell(i, 1).value
            row_data['url'] = url_uat + sheet.cell(i, 2).value

            if sheet.cell(i, 3).value.find('${phone_num}') != -1:
                if sheet.cell(i, 3).value.find('${employeeId}') != -1:
                    row_data['data'] = sheet.cell(i, 3).value.replace('${employeeId}', StartBefore().get_data_init(5))\
                                        .replace('${phone_num}', customer_phone)
                else:
                    row_data['data'] = sheet.cell(i, 3).value.replace('${phone_num}', customer_phone)

            elif sheet.cell(i, 3).value.find('${memberId}') != -1:
                if sheet.cell(i, 3).value.find('${time_now}') != -1:
                    row_data['data'] = sheet.cell(i, 3).value.replace('${time_now}', StartBefore().get_time_now()) \
                                        .replace('${memberId}', StartBefore().get_data_init(1))
                else:
                    row_data['data'] = sheet.cell(i, 3).value.replace('${memberId}', StartBefore().get_data_init(1))

            elif sheet.cell(i, 3).value.find('${storedCardId}') != -1:
                row_data['data'] = sheet.cell(i, 3).value.replace('${storedCardId}', StartBefore().get_data_init(2))
            elif sheet.cell(i, 3).value.find('${timeCardId}') != -1:
                row_data['data'] = sheet.cell(i, 3).value.replace('${timeCardId}', StartBefore().get_data_init(3))
            elif sheet.cell(i, 3).value.find('${yearCardId}') != -1:
                row_data['data'] = sheet.cell(i, 3).value.replace('${yearCardId}', StartBefore().get_data_init(4))
            elif sheet.cell(i, 3).value.find('${employeeId}') != -1:
                row_data['data'] = sheet.cell(i, 3).value.replace('${employeeId}', StartBefore().get_data_init(5))
            elif sheet.cell(i, 3).value.find('${time_now}') != -1:
                row_data['data'] = sheet.cell(i, 3).value.replace('${time_now}', StartBefore().get_time_now())
            else:
                row_data['data'] = sheet.cell(i, 3).value

            row_data['title'] = sheet.cell(i, 4).value
            row_data['http_method'] = sheet.cell(i, 5).value
            if sheet.cell(i, 6).value.find('${tk}') != -1:
                row_data['header'] = sheet.cell(i, 6).value.replace('${tk}', StartBefore().get_data_init(6))
            else:
                row_data['header'] = sheet.cell(i, 6).value
            # print(row_data['header'].find('${tk}'))
            row_data['msg'] = sheet.cell(i, 9).value
            row_data['sheet_name'] = sheet_name
            test_data.append(row_data)

        return test_data


if __name__ == '__main__':
    res = DoExcel.get_data(test_data_path, 'start_before')
    print(res)

